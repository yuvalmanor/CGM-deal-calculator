"""
Verify app calculation outputs against Excel column C (Anna TX deal).
Run from brrrr-calculator/: python scripts/verify_excel.py
"""
import subprocess
import json
import sys
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(r"C:\YuvalManorPrivate\prod\brrrr-calculator\New docs\Deal Calc CGM V2.xlsx")
NODE_RUNNER = Path("scripts/run_calc.mjs")

# Excel row numbers (1-indexed, matching Excel row display)
EXCEL_ROWS = {
    "holding_costs":      14,   # row 14: holding costs
    "all_in_cost":        15,   # row 15: cash to close (all-in, $0 PP closing costs for this deal)
    "hml_loan":           23,   # row 23: total HML loan amount
    "hml_total_debt":     28,   # row 28: total debt to HML (principal + interest + fees)
    "refi_loan":          33,   # row 33: Refi loan amount
    "cash_from_lender":   37,   # row 37: cash from Refi lender
    "net_cash_closing":   38,   # row 38: net cash at Refi closing
    "money_in_deal_hml":  39,   # row 39: money in deal (HML scenario)
    "mortgage_pi":        47,   # row 47: mortgage P+I payment
    "dscr":               52,   # row 52: DSCR
    "prop_equity":        72,   # row 72: property equity post-Refi
    "money_in_deal_cash": 55,   # row 55: money in deal (Cash scenario)
}

TOLERANCES = {
    "dscr": 0.01,
}
DEFAULT_TOLERANCE = 1.0

# Overrides for rows where the app formulas were corrected in Phase 5
# but the Excel reference sheet still uses the old definitions.
# These are the authoritative correct values — do not revert.
#
# Fix A: DSCR now includes HOA in denominator
# Fix B: HML fees paid at closing (not future debt) → lower hml_total_debt,
#         higher net_cash_closing / money_in_deal_hml
# Fix E: prop_equity = book equity (ARV − refi loan, no sale costs)
MANUAL_EXPECTED = {
    "hml_total_debt":    163980.16,
    "net_cash_closing":  19790.84,
    "money_in_deal_hml": 62861.16,
    "prop_equity":       105000.0,
    "dscr":              1.2531,
}


def read_excel_values():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    # Try common sheet name variants
    sheet_names = wb.sheetnames
    target = None
    for name in sheet_names:
        if "BRRRR" in name.upper() or "CALC" in name.upper():
            target = name
            break
    if target is None:
        print(f"Available sheets: {sheet_names}")
        print("ERROR: Could not find CALC - BRRRR sheet. Update sheet name in script.")
        sys.exit(1)
    ws = wb[target]
    values = {}
    for key, row_num in EXCEL_ROWS.items():
        cell = ws.cell(row=row_num, column=3)  # column C = index 3
        val = cell.value
        if val is None:
            print(f"WARNING: Excel row {row_num} (col C) is empty for '{key}'")
            values[key] = 0.0
        else:
            values[key] = float(val)
    return values


def run_app_calc():
    result = subprocess.run(
        ["node", str(NODE_RUNNER)],
        capture_output=True, text=True, cwd=Path.cwd()
    )
    if result.returncode != 0:
        print("Node runner failed:")
        print(result.stderr)
        sys.exit(1)
    return json.loads(result.stdout)


def compare(excel_vals, app_vals):
    passed = 0
    failed = 0
    print(f"\n{'RESULT':<25} {'EXPECTED':>15} {'APP OUTPUT':>15} {'STATUS':>10}")
    print("-" * 70)
    for key in EXCEL_ROWS:
        # Use corrected expected values where Excel retains old formulas
        expected = MANUAL_EXPECTED.get(key, excel_vals[key])
        actual = app_vals.get(key)
        if actual is None:
            print(f"{key:<25} {expected:>15.2f} {'MISSING':>15} {'FAIL':>10}")
            failed += 1
            continue
        tol = TOLERANCES.get(key, DEFAULT_TOLERANCE)
        delta = abs(expected - actual)
        ok = delta <= tol
        status = "PASS" if ok else "FAIL"
        suffix = f"  (delta: {delta:.4f})" if not ok else ""
        print(f"{key:<25} {expected:>15.4f} {actual:>15.4f} {status:>10}{suffix}")
        if ok:
            passed += 1
        else:
            failed += 1
    print()
    if failed == 0:
        print(f"ALL {passed} CHECKS PASSED")
    else:
        print(f"{failed} CHECK(S) FAILED -- fix before continuing")
        sys.exit(1)


if __name__ == "__main__":
    print("Reading Excel file...")
    excel_vals = read_excel_values()
    print("Running calculateDeal() via Node...")
    app_vals = run_app_calc()
    print("Comparing outputs...")
    compare(excel_vals, app_vals)
